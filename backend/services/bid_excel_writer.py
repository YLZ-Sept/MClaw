"""
Bid Excel Writer — appends collected bid items to desktop Excel files.
Reads items JSON from stdin, classifies and appends to:
  - G:/桌面/中标信息统计.xlsx  (win announcements)
  - G:/桌面/采购意向.xlsx       (procurement intents)

Usage: python bid_excel_writer.py <items_json>
Output: JSON with write counts per file
"""
import sys
import json
import os
import re
from datetime import datetime

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook

# ── Config ──
ZHONGBIAO_PATH = r'G:\桌面\中标信息统计.xlsx'
CAIGOU_PATH = r'G:\桌面\采购意向.xlsx'

# Keywords for classification
WIN_KEYWORDS = ['中标', '成交', '结果', '中标公告', '成交公告', '中标结果', '招标结果']
INTENT_KEYWORDS = ['采购意向', '采购需求', '采购预告', '采购计划', '需求公示', '意向公开', '采购公告', '公开招标', '招标公告', '竞争性磋商', '竞争性谈判', '询价']


def classify(title):
    """Classify item: returns 'zhongbiao' or 'caiyou'"""
    t = title or ''
    for kw in WIN_KEYWORDS:
        if kw in t:
            return 'zhongbiao'
    for kw in INTENT_KEYWORDS:
        if kw in t:
            return 'caiyou'
    # Default: if has bidding-related terms, treat as caigou
    return 'caiyou'


def read_existing_urls(filepath, url_col_index):
    """Read existing URLs from an Excel file for dedup"""
    if not os.path.exists(filepath):
        return set()
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        urls = set()
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, url_col_index).value
            if val:
                urls.add(str(val).strip())
        wb.close()
        return urls
    except Exception as e:
        print(f"[excel-writer] read existing URLs error: {e}", file=sys.stderr)
        return set()


def append_to_excel(filepath, rows, headers_count):
    """Append rows to Excel file. Returns count of actually written rows."""
    if not rows:
        return 0
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        start_row = ws.max_row + 1
        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                if j < headers_count:
                    ws.cell(start_row + i, j + 1).value = val
        wb.save(filepath)
        wb.close()
        return len(rows)
    except Exception as e:
        print(f"[excel-writer] write error for {filepath}: {e}", file=sys.stderr)
        return 0


def write_items(items):
    """Main: classify items, dedup, append to target Excel files"""
    zhongbiao_items = []
    caiyou_items = []

    for item in items:
        title = item.get('title', '')
        url = item.get('url', '')
        if not title or not url:
            continue

        category = classify(title)
        if category == 'zhongbiao':
            zhongbiao_items.append(item)
        else:
            caiyou_items.append(item)

    result = {'zhongbiao': 0, 'caiyou': 0, 'zhongbiao_skipped': 0, 'caiyou_skipped': 0}

    # ── 中标信息统计 ──
    if zhongbiao_items:
        existing_urls = read_existing_urls(ZHONGBIAO_PATH, 11)  # col 11 = 网页链接
        new_rows = []
        for item in zhongbiao_items:
            if item['url'] in existing_urls:
                result['zhongbiao_skipped'] += 1
                continue
            existing_urls.add(item['url'])
            new_rows.append([
                item.get('bid_time', ''),          # A: 中标时间
                item.get('doc_deadline', ''),       # B: 开标时间
                item.get('bid_time', ''),           # C: 投标时间
                '昆明',                              # D: 类型
                '',                                  # E: 一级行业
                item.get('source_name', ''),         # F: 招标方
                '',                                  # G: 中标公司
                item.get('title', ''),               # H: 项目名称
                item.get('purchase_requirements', ''), # I: 项目产品及服务
                item.get('amount', ''),              # J: 项目金额（万元）
                item.get('url', ''),                 # K: 网页链接
                item.get('bid_type', ''),            # L: 投标方式
                '',                                  # M: 中标单位
                '',                                  # N: 成交金额（万元）
                '',                                  # O: 备注
            ])
        if new_rows:
            result['zhongbiao'] = append_to_excel(ZHONGBIAO_PATH, new_rows, 15)

    # ── 采购意向 ──
    if caiyou_items:
        existing_urls = read_existing_urls(CAIGOU_PATH, 9)  # col 9 = 发布时间... wait, need to check column mapping

        # 采购意向.xlsx columns: A=省份, B=一级行业, C=招标方, D=项目名称, E=项目金额, F=采购需求, G=预计采购时间, H=项目地点, I=发布时间
        # No dedicated URL column in this file! We'll use 项目名称 or a combination for dedup
        # Actually let me check... the headers are: 省份, 一级行业, 招标方, 项目名称, 项目金额, 采购需求, 预计采购时间, 项目地点, 发布时间
        # No URL column. Let's use 项目名称 for dedup.
        existing_urls = set()
        if os.path.exists(CAIGOU_PATH):
            try:
                wb = load_workbook(CAIGOU_PATH)
                ws = wb.active
                for row in range(2, ws.max_row + 1):
                    val = ws.cell(row, 4).value  # D = 项目名称
                    if val:
                        existing_urls.add(str(val).strip())
                wb.close()
            except:
                pass

        new_rows = []
        for item in caiyou_items:
            # Use title for dedup since no URL column
            if item['title'] in existing_urls:
                result['caiyou_skipped'] += 1
                continue
            existing_urls.add(item['title'])
            new_rows.append([
                '昆明',                              # A: 省份
                '',                                  # B: 一级行业
                item.get('source_name', ''),         # C: 招标方
                item.get('title', ''),               # D: 项目名称
                item.get('amount', ''),              # E: 项目金额（万元）
                item.get('purchase_requirements', ''), # F: 采购需求
                item.get('doc_deadline', ''),        # G: 预计采购时间
                '',                                  # H: 项目地点
                item.get('fetch_time', ''),          # I: 发布时间
            ])
        if new_rows:
            result['caiyou'] = append_to_excel(CAIGOU_PATH, new_rows, 9)

    return result


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'Usage: python bid_excel_writer.py <items_json>'}))
        sys.exit(1)

    try:
        items = json.loads(sys.argv[1])
        if not isinstance(items, list):
            items = [items]
    except json.JSONDecodeError as e:
        print(json.dumps({'error': f'JSON parse error: {e}'}, ensure_ascii=False))
        sys.exit(1)

    result = write_items(items)
    print(json.dumps(result, ensure_ascii=False))
