"""导出 — JSON 清洗、字段映射、输出 xlsx"""
import os
import sys
import json
import argparse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from config import DATA_DIR, OUTPUT_DIR, TAB_CONFIG


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def clean_text(text):
    """清洗文本：去 HTML 标签、去多余空格"""
    import re
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", "", str(text))
    text = re.sub(r"\s+", " ", text).strip()
    return text


def format_date(text):
    """标准化日期格式"""
    if not text:
        return ""
    text = str(text).strip()
    # 处理 "今天"、"1小时前更新" 等相对时间
    if "今天" in text:
        return datetime.now().strftime("%Y-%m-%d")
    if "前" in text:
        return datetime.now().strftime("%Y-%m-%d")
    # 标准化 YYYY-MM-DD 或 YYYY/MM/DD
    import re
    m = re.search(r"(\d{4})[/.-](\d{1,2})[/.-](\d{1,2})", text)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    return text[:10]


def extract_amount_value(text):
    """从文本中提取纯数字金额（万元）"""
    import re
    if not text:
        return ""
    text = str(text)
    # 处理 "260.00万元" "26.00万元" "300.00万元"
    m = re.search(r"(\d+\.?\d*)\s*万", text)
    if m:
        return float(m.group(1))
    # 处理纯数字
    m = re.search(r"(\d+\.?\d*)", text)
    if m:
        val = float(m.group(1))
        if val > 10000:  # 可能是元为单位
            return val / 10000
        return val
    return ""


def extract_field_from_text(text, *labels):
    """从文本中提取标签后的值"""
    import re
    if not text:
        return ""
    for label in labels:
        # 匹配 "标签：值" 或 "标签\t值" (不跨行)
        patterns = [
            rf"{label}[：:]\s*([^\n]{{2,80}})",
            rf"{label}\t([^\n]{{2,80}})",
        ]
        for pat in patterns:
            m = re.search(pat, text)
            if m:
                val = m.group(1).strip()
                invalid = ["<", "font", "span", "div", "公告", "采购", "招标", "项目"]
                if val and len(val) > 1 and not any(val.startswith(w) for w in invalid):
                    return val[:200]
    return ""


def map_zhongbiao(item):
    """将采集的原始数据映射到中标信息 Excel 列"""
    title = clean_text(item.get("title", ""))
    content = clean_text(item.get("content", "") or item.get("_fullContent", ""))
    area = item.get("areaName", "") or item.get("area", "")
    pub_type = item.get("type", "")

    # 时间字段
    update_date = format_date(item.get("updateDate", ""))
    reg_deadline = clean_text(item.get("_regDeadline", "") or item.get("registrationDeadline", ""))
    bid_deadline = clean_text(item.get("_bidDeadline", "") or item.get("tenderEndTime", ""))
    bid_time = format_date(item.get("_bidDeadline", "") or item.get("tenderEndTime", ""))

    # 金额
    amount_raw = item.get("_amount", "") or item.get("amountUnit", "") or item.get("extractBudget", "")
    amount = extract_amount_value(amount_raw)
    deal_amount_raw = item.get("_dealAmount", "")
    deal_amount = extract_amount_value(deal_amount_raw)

    # 合并所有文本用于字段提取
    full_text = f"{content} {clean_text(item.get('_fullContent', '') or '')}"
    detail_text = clean_text(item.get("_fullContent", "") or "")

    # 单位和链接
    raw_bidder = item.get("招标单位", "") or item.get("zhaoBiaoUnit", "")
    bidder = extract_field_from_text(detail_text, "采购单位", "招标单位", "采购人") or raw_bidder
    # 清洗：只取第一行有效内容
    if bidder and "\n" in str(bidder):
        bidder = str(bidder).split("\n")[0].strip()

    raw_agent = item.get("代理单位", "") or item.get("agentUnit", "")
    agent = extract_field_from_text(detail_text, "代理机构名称", "代理单位", "采购代理机构") or raw_agent
    if agent and "\n" in str(agent):
        agent = str(agent).split("\n")[0].strip()

    raw_winner = item.get("中标单位", "") or item.get("zhongBiaoUnit", "")
    winner = extract_field_from_text(full_text, "中标单位", "中标供应商", "供应商名称", "成交供应商") or raw_winner
    if winner and "\n" in str(winner):
        winner = str(winner).split("\n")[0].strip()
    content_id = item.get("contentId", "")
    area_id = item.get("areaId", "29")
    link = f"https://qiye.qianlima.com/new_qd_yfbsite/#/infoCenter/infoDetail/{content_id}/{area_id}/zhongbiao?fromPage=searchPage" if content_id else ""

    # 采购方式
    procurement_method = (
        item.get("采购方式", "") or
        extract_field_from_text(detail_text, "采购方式") or
        item.get("biddingType", "")
    )
    # 投标截止时间
    bid_time_from_detail = (
        item.get("投标截止时间", "") or
        extract_field_from_text(detail_text, "投标截止时间", "响应文件提交截止时间")
    )
    if bid_time_from_detail:
        bid_time = format_date(bid_time_from_detail)

    # 行业分类
    from config import classify_industry
    industry = classify_industry(title, content)

    row = {
        "招标时间": update_date,
        "报名时间": reg_deadline if reg_deadline != "-" else "",
        "投标时间": bid_time,
        "区域": area,
        "一级行业": industry,
        "招标人": bidder,
        "招标公司": agent,
        "项目名称": title,
        "项目产品（服务）": content[:300] if content else "",
        "项目金额（万元）": amount,
        "网页链接": link,
        "招投标方式": procurement_method,
        "中标单位": winner,
        "成交金额（万元）": deal_amount,
        "备注": pub_type,
    }
    return row


def map_caiyi(item):
    """将采集的原始数据映射到采购意向 Excel 列"""
    title = clean_text(item.get("title", ""))
    if item.get("newTitle"):
        title = clean_text(item["newTitle"])

    area = item.get("area", "") or item.get("areaName", "")
    bidder = clean_text(item.get("zhaoBiaoUnit", "") or item.get("zhaoBiaoUnitComplate", ""))
    amount = extract_amount_value(item.get("budgetAmount", "") or item.get("budget", ""))
    purchase_survey = clean_text(item.get("purchaseSurvey", "") or item.get("content", ""))

    # 时间
    release_time = item.get("releaseTime", "") or item.get("releaseTimeYyyymmdd", "")
    release_time = format_date(release_time)
    estimate_time = item.get("estimateTime", "") or item.get("estimateTimeYyyymm", "")

    # 行业分类
    from config import classify_industry
    industry = classify_industry(title, purchase_survey)

    # 链接
    item_id = item.get("id", "")
    relation_id = item.get("relationId", "")
    area_id = item.get("areaId", "29")
    if relation_id:
        link = (f"https://qiye.qianlima.com/new_qd_yfbsite/#/infoCenter/infoDetail/"
                f"{relation_id}/{area_id}/caigou?purchaseId={item_id}&fromPage=searchPage")
    else:
        link = ""

    row = {
        "区域": area,
        "一级行业": industry,
        "招标人": bidder,
        "项目名称": title,
        "项目金额（万元）": amount,
        "采购需求": purchase_survey[:500] if purchase_survey else "",
        "预计采购时间": estimate_time,
        "项目地点": area,
        "发布时间": release_time,
    }
    return row


def style_sheet(ws, headers):
    """统一设置 Excel 样式"""
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="微软雅黑", size=10)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 设置列宽
    col_widths = {"项目名称": 50, "项目产品（服务）": 40, "采购需求": 50,
                   "网页链接": 35, "备注": 20, "区域": 12, "一级行业": 15}
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = \
            col_widths.get(header, 18)

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions


def export(data_file=None):
    """导出到 xlsx"""
    ensure_dir(OUTPUT_DIR)

    # 查找最新的数据文件
    if not data_file:
        batches = sorted(os.listdir(DATA_DIR), reverse=True)
        if not batches:
            print("没有找到采集数据，请先运行: python scrape.py")
            sys.exit(1)
        data_file = os.path.join(DATA_DIR, batches[0], "_full_data.json")

    if not os.path.exists(data_file):
        # 尝试 _raw_list.json
        data_dir = os.path.dirname(data_file)
        raw_file = os.path.join(data_dir, "_raw_list.json")
        if os.path.exists(raw_file):
            data_file = raw_file
        else:
            print(f"数据文件不存在: {data_file}")
            sys.exit(1)

    print(f"读取数据: {data_file}")
    with open(data_file, "r", encoding="utf-8") as f:
        all_data = json.load(f)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 导出中标信息
    zhongbiao_items = all_data.get("zhongbiao", [])
    if zhongbiao_items:
        print(f"导出中标信息: {len(zhongbiao_items)} 条")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "中标信息"

        headers = TAB_CONFIG["zhongbiao"]["headers"]
        style_sheet(ws, headers)

        for row_idx, item in enumerate(zhongbiao_items, 2):
            mapped = map_zhongbiao(item)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=mapped.get(header, ""))
                cell.font = Font(name="微软雅黑", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"),
                )

        filename = f"中标信息统计_{timestamp}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        wb.save(filepath)
        print(f"已保存: {filepath}")

    # 导出采购意向
    caiyi_items = all_data.get("caiyi", [])
    if caiyi_items:
        print(f"导出采购意向: {len(caiyi_items)} 条")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "采购意向"

        headers = TAB_CONFIG["caiyi"]["headers"]
        style_sheet(ws, headers)

        for row_idx, item in enumerate(caiyi_items, 2):
            mapped = map_caiyi(item)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=mapped.get(header, ""))
                cell.font = Font(name="微软雅黑", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"),
                )

        filename = f"采购意向统计_{timestamp}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        wb.save(filepath)
        print(f"已保存: {filepath}")

    print("\n导出完成!")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="导出到 Excel")
    parser.add_argument("--data", help="JSON 数据文件路径")
    args = parser.parse_args()
    export(data_file=args.data)
